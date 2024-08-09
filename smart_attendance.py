import os
import cv2
import face_recognition
import openpyxl
from datetime import datetime
import dlib
from scipy.spatial import distance

# Initialize the Excel workbook
def init_excel(file_path):
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Visitors'
        sheet.append(['ID', 'Name', 'Register Number', 'Date' , 'Time' , 'Status'])  # Updated header
        workbook.save(file_path)

# Add a record to the Excel file with separate columns for date and time, and a report column
def log_visitor(file_path, name, register_number, known_visitors):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['Visitors']
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")
    
    # Check if the person has already been logged in this session
    if name in known_visitors:
        print(f"Attendance already taken for {name}")
        return "Attendance already taken " + name
    else:
        known_visitors.add(name)
        next_id = sheet.max_row + 1 if sheet.max_row > 1 else 1
        sheet.append([next_id, name, register_number, date, time, "Present"])  # Updated to include date, time, and report
        workbook.save(file_path)
        print(f"Present for {name} (Register Number: {register_number})")
        return "Present " + name


# Load known faces and names from a specific directory
known_face_encodings = []
known_face_names = []

def load_known_faces(directory):
    for filename in os.listdir(directory):
        if filename.endswith(".jpg") or filename.endswith(".png"):  # Add other file formats if needed
            filepath = os.path.join(directory, filename)
            image = face_recognition.load_image_file(filepath)
            encodings = face_recognition.face_encodings(image)
            if encodings:
                known_face_encodings.append(encodings[0])
                # Assuming filename format is name_registernumber.jpg/png
                name, register_number = os.path.splitext(filename)[0].split('_')
                known_face_names.append((name, register_number))

# Calculate Eye Aspect Ratio (EAR)
def calculate_ear(eye):
    A = distance.euclidean(eye[1], eye[5])
    B = distance.euclidean(eye[2], eye[4])
    C = distance.euclidean(eye[0], eye[3])
    ear = (A + B) / (2.0 * C)
    return ear

# Main function to run the attendance system
def run_attendance_system():
    # Create a new Excel file with timestamp for uniqueness
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    file_path = f'attendance_{current_time}.xlsx'
    init_excel(file_path)

    # Initialize known faces and names
    known_faces_directory = r'C:\Users\JEEVAN\Projects\Smart Attendance System\known_faces'
    load_known_faces(known_faces_directory)

    # Start video capture
    video_capture = cv2.VideoCapture(0)

    # Set to track known visitors in current session
    known_visitors = set()

    # Initialize dlib's face detector and shape predictor for blink detection
    predictor_path = "C:\\Users\\JEEVAN\\Projects\\Smart Attendance System\\shape_predictor_68_face_landmarks.dat"  # Adjust the path if needed
    
    
    detector = dlib.get_frontal_face_detector()
    predictor = dlib.shape_predictor(predictor_path)

    EYE_AR_THRESH = 0.25  # Adjusted threshold for EAR
    EYE_AR_CONSEC_FRAMES = 2
    blink_counter = 0
    blink_detected = False

    while True:
        ret, frame = video_capture.read()
        if not ret:
            print("Failed to grab frame")
            break

        # Convert the frame to RGB
        rgb_frame = frame[:, :, ::-1]

        # Find all face locations and face encodings in the current frame
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        for face_encoding, face_location in zip(face_encodings, face_locations):
            # See if the face matches any known faces
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"
            register_number = ""

            if True in matches:
                first_match_index = matches.index(True)
                name, register_number = known_face_names[first_match_index]
            else:
                print("Unknown person detected!")

            # Detect faces and find landmarks for blink detection
            dlib_rects = detector(rgb_frame, 0)
            for dlib_rect in dlib_rects:
                shape = predictor(rgb_frame, dlib_rect)
                shape = face_recognition.face_landmarks(rgb_frame, [face_location])[0]

                left_eye = shape['left_eye']
                right_eye = shape['right_eye']

                left_ear = calculate_ear(left_eye)
                right_ear = calculate_ear(right_eye)
                ear = (left_ear + right_ear) / 2.0


                if ear < EYE_AR_THRESH:
                    blink_counter += 1
                else:
                    if blink_counter >= EYE_AR_CONSEC_FRAMES:
                        blink_detected = True
                    blink_counter = 0

            if blink_detected:
                # Log the visitor in the Excel file and display status on screen
                status = log_visitor(file_path, name, register_number, known_visitors)
            else:
                status = "Face not Clear"

            # Draw a box around the face
            top, right, bottom, left = face_location
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
            cv2.putText(frame, status, (left + 6, bottom - 6), cv2.FONT_HERSHEY_DUPLEX, 0.5, (255, 255, 255), 1)

        # Display the resulting frame
        cv2.imshow('Video', frame)

        # Press 'q' to quit
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release the video capture and close windows
    video_capture.release()
    cv2.destroyAllWindows()

# Run the attendance system
if __name__ == "__main__":
    run_attendance_system()